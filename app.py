import streamlit as st
import google.generativeai as genai
import pandas as pd
import json
from PIL import Image
import unicodedata
import io
import os

# Importaciones para extracción avanzada de enlaces web (Caso Bahco)
import requests
from bs4 import BeautifulSoup

# Importaciones nativas para diseño avanzado de Excel y PDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from fpdf import FPDF

# Configuración de la página web (Tu marca personal de software)
st.set_page_config(page_title="Cotizador IA EHP", layout="wide")

# AUTODETECCIÓN MAESTRA DEL LOGO CORPORATIVO EN EL SERVIDOR
# Se conserva en memoria EXCLUSIVAMENTE para inyectarlo en los entregables PDF y Excel del cliente
logo_bytes = None
for ext in ["png", "jpg", "jpeg"]:
    if os.path.exists(f"logo.{ext}"):
        with open(f"logo.{ext}", "rb") as f:
            logo_bytes = f.read()
        break

# TÍTULO PRINCIPAL DE LA INTERFAZ WEB NATIVA
st.title("Cotizador IA EHP 🔧")

# Inicializar estados de memoria estables para evitar borrados al interactuar en la pantalla
if 'df_resultado' not in st.session_state:
    st.session_state['df_resultado'] = None
    st.session_state['nombre_cliente_s'] = ""
    st.session_state['empresa_cliente_s'] = ""
    st.session_state['numero_folio_s'] = ""
    st.session_state['condicion_pago_s'] = "CONTADO"
    st.session_state['vendedor_s'] = "Enrique Hernández P."

# Lista de palabras vacías en español para limpiar búsquedas del catálogo
STOP_WORDS = {'de', 'para', 'con', 'un', 'una', 'el', 'la', 'los', 'las', 'del', 'al', 'en', 'y', 'por', 'sobre', 'kit', 'juego', 'set'}

def normalizar_texto(texto):
    if pd.isna(texto):
        return ""
    texto = str(texto).lower().strip()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')
    for char in ['( ', ' )', '(', ')', '-', ',', '.', '/', '"', "'", '+']:
        texto = texto.replace(char, ' ')
    return texto

def limpiar_plurales(texto):
    palabras = texto.split()
    limpias = []
    for p in palabras:
        if len(p) > 3 and p.endswith('s'):
            limpias.append(p[:-1])
        else:
            limpias.append(p)
    return " ".join(limpias)

def limpiar_precio(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(int(valor))
    val_str = str(valor).strip().replace("$", "").replace(" ", "")
    if not val_str:
        return 0.0
    if "," in val_str and val_str.count(",") == 1 and len(val_str.split(",")[-1]) <= 2:
        val_str = val_str.split(",")[0]
    if "." in val_str and val_str.count(".") == 1 and len(val_str.split(".")[-1]) <= 2:
        val_str = val_str.split(".")[0]
    val_str = val_str.replace(".", "").replace(",", "")
    try:
        return float(int(val_str))
    except:
        return 0.0

def leer_csv_tolerante(ruta_archivo):
    if not os.path.exists(ruta_archivo):
        return None
    encodings_a_probar = ['utf-8', 'latin1', 'iso-8859-1', 'utf-8-sig', 'cp1252']
    delimitadores_a_probar = [';', ',']
    for enc in encodings_a_probar:
        for sep in delimitadores_a_probar:
            try:
                with open(ruta_archivo, 'r', encoding=enc) as f:
                    lineas = [f.readline() for _ in range(15)]
                fila_cabecera_idx = 0
                for idx, line in enumerate(lineas):
                    line_low = line.lower()
                    coincidencias = sum(1 for kw in ['cod', 'desc', 'prec', 'mar', 'art', 'vta', 'neto', 'prod', 'clien'] if kw in line_low)
                    if coincidencias >= 2:
                        fila_cabecera_idx = idx
                        break
                df_res = pd.read_csv(ruta_archivo, sep=sep, encoding=enc, skiprows=fila_cabecera_idx)
                df_res = df_res.dropna(how='all', axis=1)
                df_res = df_res.dropna(how='all', axis=0)
                if df_res.shape[1] >= 2:
                    return df_res
            except:
                continue
    return None

def leer_equivalencias(ruta_archivo="equivalencias.csv"):
    if not os.path.exists(ruta_archivo):
        return {}
    encodings = ['utf-8', 'latin1', 'iso-8859-1', 'utf-8-sig', 'cp1252']
    delimitadores = [';', ',']
    for enc in encodings:
        for sep in delimitadores:
            try:
                df_eq = pd.read_csv(ruta_archivo, sep=sep, encoding=enc)
                df_eq.columns = [str(c).strip().lower() for c in df_eq.columns]
                col_term = next((c for c in df_eq.columns if 'term' in c or 'clien' in c or 'busq' in c), df_eq.columns[0])
                col_cod = next((c for c in df_eq.columns if 'cod' in c or 'cat' in c), df_eq.columns[1] if len(df_eq.columns) > 1 else df_eq.columns[0])
                dict_eq = {}
                for _, r in df_eq.iterrows():
                    key_cliente = normalizar_texto(r[col_term])
                    val_codigo = str(r[col_cod]).strip()
                    if key_cliente and val_codigo:
                        dict_eq[key_cliente] = val_codigo
                return dict_eq
            except:
                continue
    return {}

# RASPADOR NATIVO DE TEXTO WEB: Extrae el contenido limpio de URLs oficiales (Caso Bahco)
def extraer_texto_de_url(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        r = requests.get(url, headers=headers, timeout=10)
        if r.status_code == 200:
            soup = BeautifulSoup(r.text, 'html.parser')
            for script in soup(["script", "style", "header", "footer", "nav"]):
                script.extract()
            lines = (line.strip() for line in soup.get_text().splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            return " ".join(chunk for chunk in chunks if chunk)
    except:
        pass
    return ""

def clean_pdf_str(text):
    if pd.isna(text):
        return ""
    return str(text).encode('latin-1', 'replace').decode('latin-1')

# FUNCIÓN: Generación de PDF Comercial Oficial
def generar_pdf_comercial(df_cotiz, cliente, empresa, nro_cotiz, total_neto, iva, total_bruto, logo_bytes=None, condicion_pago="CONTADO", vendedor="Enrique Hernández P."):
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    if logo_bytes:
        try: pdf.image(io.BytesIO(logo_bytes), x=10, y=10, w=38)
        except: pass
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(0, 5, clean_pdf_str(f"Fecha: {pd.Timestamp.now().strftime('%d-%m-%Y')}"), ln=True, align="R")
    pdf.ln(6)
    pdf.set_font("helvetica", "B", 16)
    pdf.set_text_color(31, 73, 125)
    pdf.cell(0, 10, clean_pdf_str(f"COTIZACIÓN N° {nro_cotiz}"), ln=True, align="C")
    pdf.ln(4)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(0, 5, clean_pdf_str("76.834.968-1"), ln=True)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, clean_pdf_str("Chopin 2848. San Joaquín. Santiago"), ln=True)
    pdf.ln(3)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(0, 5, clean_pdf_str(f"Sr(a).: {cliente if cliente else 'No especificado'}"), ln=True)
    pdf.cell(0, 5, clean_pdf_str(f"Empresa: {empresa if empresa else 'No especificada'}"), ln=True)
    pdf.ln(3)
    pdf.set_font("helvetica", "", 10)
    pdf.cell(0, 5, clean_pdf_str("En atención a su gentil solicitud de cotización, tenemos el agrado de hacer llegar a usted nuestra propuesta:"), ln=True)
    pdf.ln(4)
    pdf.set_font("helvetica", "B", 9)
    pdf.set_fill_color(54, 95, 145)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(24, 8, clean_pdf_str("CÓDIGO"), border=1, align="C", fill=True)
    pdf.cell(24, 8, clean_pdf_str("MARCA"), border=1, align="C", fill=True)
    pdf.cell(72, 8, clean_pdf_str("DESCRIPCIÓN"), border=1, align="C", fill=True)
    pdf.cell(28, 8, clean_pdf_str("P. UNIT NETO"), border=1, align="C", fill=True)
    pdf.cell(14, 8, clean_pdf_str("CANT"), border=1, align="C", fill=True)
    pdf.cell(28, 8, clean_pdf_str("TOTAL NETO"), border=1, align="C", fill=True)
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("helvetica", "", 9)
    for _, fila in df_cotiz.iterrows():
        cod = str(fila["Código"])
        marca = str(fila["Marca"])
        desc = str(fila["Descripción Catálogo"]).replace("⚠️ (Match sugerido) ", "")
        p_unit = f"${float(fila['Precio Final (Neto)']):,.0f}"
        cant = str(int(fila["Cantidad"]))
        total = f"${float(fila['Total Neto']):,.0f}"
        if len(desc) > 42: desc = desc[:39] + "..."
        pdf.cell(24, 7, clean_pdf_str(cod), border=1, align="C")
        pdf.cell(24, 7, clean_pdf_str(marca), border=1, align="C")
        pdf.cell(72, 7, clean_pdf_str(desc), border=1, align="L")
        pdf.cell(28, 7, clean_pdf_str(p_unit), border=1, align="R")
        pdf.cell(14, 7, clean_pdf_str(cant), border=1, align="C")
        pdf.cell(28, 7, clean_pdf_str(total), border=1, align="R")
        pdf.ln(7)
    pdf.ln(5)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(120, 6, clean_pdf_str("Observaciones:"), border=0)
    pdf.cell(42, 6, clean_pdf_str("SUBTOTAL"), border=1, align="R")
    pdf.cell(28, 6, clean_pdf_str(f"${total_neto:,.0f}"), border=1, align="R")
    pdf.ln(6)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(120, 6, clean_pdf_str("1: Plazo de entrega por confirmar"), border=0)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(42, 6, clean_pdf_str("IVA"), border=1, align="R")
    pdf.cell(28, 6, clean_pdf_str(f"${iva:,.0f}"), border=1, align="R")
    pdf.ln(6)
    pdf.set_font("helvetica", "", 9)
    pdf.cell(120, 6, clean_pdf_str("2. Validez de cotización: 7 días"), border=0)
    pdf.set_font("helvetica", "B", 10)
    pdf.set_fill_color(233, 237, 244)
    pdf.cell(42, 6, clean_pdf_str("TOTAL"), border=1, align="R")
    pdf.cell(28, 6, clean_pdf_str(f"${total_bruto:,.0f}"), border=1, align="R", fill=True)
    pdf.ln(8)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(120, 5, clean_pdf_str(f"Condiciones de Pago: {condicion_pago.upper()}"), ln=0)
    pdf.set_font("helvetica", "BI", 11)
    pdf.cell(70, 5, clean_pdf_str(f"{vendedor}"), ln=1, align="R")
    pdf.set_font("helvetica", "", 10)
    pdf.cell(120, 5, clean_pdf_str("A la espera de sus comentarios, le saluda atentamente :"), ln=0)
    pdf.set_font("helvetica", "B", 10)
    pdf.cell(70, 5, clean_pdf_str("VGM SpA"), ln=1, align="R")
    return pdf.output()

# FUNCIÓN: Generación de Excel Comercial Oficial
def generar_excel_comercial(df_cotiz, cliente, empresa, nro_cotiz, total_neto, iva, total_bruto, logo_bytes=None, dict_imagenes=None, condicion_pago="CONTADO", vendedor="Enrique Hernández P."):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"
    ws.views.sheetView[0].showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0; ws.page_margins.right = 0; ws.page_margins.top = 0; ws.page_margins.bottom = 0
    font_titulo = Font(name="Arial", size=16, bold=True, color="1F497D")
    font_cabecera_tabla = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    font_negrita = Font(name="Arial", size=10, bold=True)
    font_normal = Font(name="Arial", size=10)
    font_firma = Font(name="Arial", size=11, bold=True, italic=True)
    fill_azul_header = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
    fill_totales = PatternFill(start_color="E9EDF4", end_color="E9EDF4", fill_type="solid")
    borde_delgado = Border(left=Side(style='thin', color='BFBFBF'), right=Side(style='thin', color='BFBFBF'), top=Side(style='thin', color='BFBFBF'), bottom=Side(style='thin', color='BFBFBF'))
    ws["G1"] = f"Fecha: {pd.Timestamp.now().strftime('%d-%m-%Y')}"
    ws["G1"].font = font_negrita
    ws["G1"].alignment = Alignment(horizontal="right")
    if logo_bytes:
        try:
            img_stream = io.BytesIO(logo_bytes)
            img = OpenpyxlImage(img_stream)
            img.width = 130; img.height = 50
            ws.add_image(img, 'A1')
        except: pass
    ws.merge_cells("A3:G3")
    celda_tit = ws["A3"]; celda_tit.value = f"COTIZACIÓN N°{nro_cotiz}"; celda_tit.font = font_titulo
    celda_tit.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 32
    ws["A4"] = "76.834.968-1"; ws["A4"].font = font_negrita
    ws["A5"] = "Chopin 2848. San Joaquín. Santiago"; ws["A5"].font = font_normal
    ws["A6"] = f"Sr(a).: {cliente if cliente else 'No especificado'}"; ws["A6"].font = font_negrita
    ws["A7"] = f"Empresa: {empresa if empresa else 'No especificada'}"; ws["A7"].font = font_negrita
    ws["A8"] = "En atención a su gentil solicitud de cotización, tenemos el agrado de hacer llegar a usted nuestra propuesta:"; ws["A8"].font = font_normal
    
    titulos_columnas = ["CODIGO", "MARCA", "DESCRIPCIÓN", "PRECIO UNITARIO NETO", "CANTIDAD", "PRECIO UNITARIO TOTAL", "IMAGEN REFERENCIAL"]
    fila_tabla_inicio = 10
    for col_idx, texto_col in enumerate(titulos_columnas, 1):
        celda = ws.cell(row=fila_tabla_inicio, column=col_idx, value=texto_col)
        celda.font = font_cabecera_tabla; celda.fill = fill_azul_header
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); celda.border = borde_delgado
    ws.row_dimensions[fila_tabla_inicio].height = 28
    
    fila_actual = fila_tabla_inicio + 1
    for _, fila in df_cotiz.iterrows():
        cod_original = str(fila["Código"])
        cod_limpio = cod_original.strip().lower()
        ws.cell(row=fila_actual, column=1, value=cod_original).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=fila_actual, column=2, value=str(fila["Marca"])).alignment = Alignment(horizontal="center", vertical="center")
        desc_excel = str(fila["Descripción Catálogo"]).replace("⚠️ (Match sugerido) ", "")
        ws.cell(row=fila_actual, column=3, value=desc_excel).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_neto = ws.cell(row=fila_actual, column=4, value=float(fila["Precio Final (Neto)"]))
        c_neto.number_format = '$#,##0'; c_neto.alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=fila_actual, column=5, value=int(fila["Cantidad"])).alignment = Alignment(horizontal="center", vertical="center")
        c_total = ws.cell(row=fila_actual, column=6, value=float(fila["Total Neto"]))
        c_total.number_format = '$#,##0'; c_total.alignment = Alignment(horizontal="right", vertical="center")
        
        if dict_imagenes and cod_limpio in dict_imagenes:
            try:
                img_prod_stream = io.BytesIO(dict_imagenes[cod_limpio])
                img_excel = OpenpyxlImage(img_prod_stream)
                img_excel.width = 115; img_excel.height = 80
                ws.add_image(img_excel, f"G{fila_actual}")
                ws.row_dimensions[fila_actual].height = 65
            except: ws.row_dimensions[fila_actual].height = 22
        else: ws.row_dimensions[fila_actual].height = 22
        for c_idx in range(1, 8):
            ws.cell(row=fila_actual, column=c_idx).border = borde_delgado; ws.cell(row=fila_actual, column=c_idx).font = font_normal
        fila_actual += 1
        
    ws.cell(row=fila_actual, column=1, value="Observaciones:").font = font_negrita
    celda_lbl_sub = ws.cell(row=fila_actual, column=5, value="SUBTOTAL"); celda_lbl_sub.font = font_negrita; celda_lbl_sub.alignment = Alignment(horizontal="right", vertical="center"); celda_lbl_sub.border = borde_delgado
    celda_val_sub = ws.cell(row=fila_actual, column=6, value=total_neto); celda_val_sub.font = font_negrita; celda_val_sub.number_format = '$#,##0'; celda_val_sub.alignment = Alignment(horizontal="right", vertical="center"); celda_val_sub.border = borde_delgado
    fila_actual += 1
    celda_lbl_iva = ws.cell(row=fila_actual, column=5, value="IVA"); celda_lbl_iva.font = font_negrita; celda_lbl_iva.alignment = Alignment(horizontal="right", vertical="center"); celda_lbl_iva.border = borde_delgado
    celda_val_iva = ws.cell(row=fila_actual, column=6, value=iva); celda_val_iva.font = font_negrita; celda_val_iva.number_format = '$#,##0'; celda_val_iva.alignment = Alignment(horizontal="right", vertical="center"); celda_val_iva.border = borde_delgado
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value="Condiciones de Venta:").font = font_negrita
    celda_lbl_tot = ws.cell(row=fila_actual, column=5, value="TOTAL"); celda_lbl_tot.font = font_negrita; celda_lbl_tot.alignment = Alignment(horizontal="right", vertical="center"); celda_lbl_tot.border = borde_delgado
    celda_val_tot = ws.cell(row=fila_actual, column=6, value=total_bruto); celda_val_tot.font = font_negrita; celda_val_tot.fill = fill_totales; celda_val_tot.number_format = '$#,##0'; celda_val_tot.alignment = Alignment(horizontal="right", vertical="center"); celda_val_tot.border = borde_delgado
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value="1: Plazo de entrega por confirmar").font = font_normal
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value="2. Validez de cotización: 7 días").font = font_normal
    ws.cell(row=fila_actual, column=7, value=vendedor).font = font_firma; ws.cell(row=fila_actual, column=7).alignment = Alignment(horizontal="right")
    fila_actual += 1
    ws.cell(row=fila_actual, column=1, value=f"Condiciones de Pago: {condicion_pago.upper()}").font = font_negrita
    ws.cell(row=fila_actual, column=7, value="VGM SpA").font = font_negrita; ws.cell(row=fila_actual, column=7).alignment = Alignment(horizontal="right")
    
    ws.column_dimensions['A'].width = 14; ws.column_dimensions['B'].width = 14; ws.column_dimensions['C'].width = 46; ws.column_dimensions['D'].width = 24; ws.column_dimensions['E'].width = 12; ws.column_dimensions['F'].width = 24; ws.column_dimensions['G'].width = 24
    wb.save(output)
    return output.getvalue()

# Interfaz de Usuario Lateral NATIVA (Sidebar)
with st.sidebar:
    st.subheader("🔑 Motor de Inteligencia")
    api_key = None
    if "GEMINI_API_KEY" in st.secrets and st.secrets["GEMINI_API_KEY"].strip():
        api_key = st.secrets["GEMINI_API_KEY"]
        st.success("🔑 Motor Gemini autenticado.")
    else:
        api_key = st.text_input("Ingresa tu Gemini API Key:", type="password")
        if api_key.strip(): st.success("🔑 Motor Gemini autenticado.")
            
    st.markdown("---")
    st.subheader("💼 Datos de la Cotización")
    nombre_cliente = st.text_input("Nombre del Cliente:", placeholder="Ej: José Mendoza")
    empresa_cliente = st.text_input("Empresa / Entidad:", placeholder="Ej: Llantas del Pacífico")
    numero_folio = st.text_input("Número de Cotización:", value="EHP-TSA-2026")
    condicion_pago_input = st.text_input("Condición de Pago:", value="CONTADO")
    vendedor_input = st.text_input("Vendedor:", value="Enrique Hernández P.")
    
    st.markdown("---")
    st.subheader("🔀 Configuración de Precios")
    # SELECTOR MAESTRO DE MODO COMERCIAL
    modo_operacion = st.selectbox(
        "Modo de Operación / Origen:",
        ["Catálogo Interno (Precio Lista)", "Costo Proveedor + Margen (Pantallazos/Links)"]
    )
    
    if modo_operacion == "Catálogo Interno (Precio Lista)":
        descuento_aplicar = st.number_input("Descuento a aplicar (%)", min_value=0, max_value=100, value=0, step=1)
        margen_objetivo = 0
    else:
        margen_objetivo = st.number_input("Margen Objetivo (%)", min_value=0, max_value=500, value=30, step=5)
        descuento_aplicar = 0
        
    precio_manual_input = st.text_input("Precio Neto Fijo Alternativo (Opcional):", placeholder="Ej: 500000")

# Renderizado de Paneles Centrales dependiendo del Modo de Operación
if modo_operacion == "Catálogo Interno (Precio Lista)":
    st.subheader("1. Carga la Solicitud del Cliente (Elige el formato)")
    tab_imagen, tab_texto, tab_pdf = st.tabs(["📸 Pantallazo / Imagen", "✍️ Copiar-Pegar Texto", "📄 Archivo PDF"])
    imagen_pedido, texto_pedido, pdf_pedido = None, "", None
    with tab_imagen: imagen_pedido = st.file_uploader("Selecciona la imagen", type=["png", "jpg", "jpeg"])
    with tab_texto: texto_pedido = st.text_area("Pega el texto aquí:")
    with tab_pdf: pdf_pedido = st.file_uploader("Sube el PDF", type=["pdf"])
    
    df_catalogo = leer_csv_tolerante("lista_vigente.csv")
    if df_catalogo is None:
        st.error("❌ No se encontró el catálogo de precios 'lista_vigente.csv' en GitHub.")
        st.stop()
        
    input_listo = False
    contenido_para_gemini = []
    prompt_extraccion = """
    Analiza detalladamente esta solicitud de pedido (puede ser imagen, texto o PDF). Extrae cada producto solicitado y su cantidad precisa. Genera de 4 a 6 sinónimos o términos técnicos para ampliar la búsqueda.
    Devuelve ÚNICAMENTE un JSON puro con esta estructura, sin bloques markdown:
    {
        "productos": [
            {"busqueda": "Nombre original o término identificado", "cantidad": 1, "sinonimos": ["t1", "t2"]}
        ]
    }
    """
    if imagen_pedido:
        input_listo = True
        contenido_para_gemini = [prompt_extraccion, Image.open(imagen_pedido)]
    elif texto_pedido.strip():
        input_listo = True
        contenido_para_gemini = [prompt_extraccion + "\n\nTEXTO:\n" + texto_pedido]
    elif pdf_pedido:
        input_listo = True
        contenido_para_gemini = [prompt_extraccion, {"mime_type": "application/pdf", "data": pdf_pedido.read()}]

else:
    # MODO: Costo Proveedor + Margen
    st.subheader("1. Carga los Pantallazos o Links del Proveedor (Camino B)")
    tab_imagenes_prov, tab_links_prov = st.tabs(["📸 Múltiples Pantallazos / Capturas", "🔗 Enlaces / Fichas de Fabricantes"])
    
    imagenes_prov = []
    with tab_imagenes_prov: 
        imagenes_prov = st.file_uploader("Sube uno o varios pantallazos de portales de proveedores juntos:", type=["png", "jpg", "jpeg"], accept_multiple_files=True)
    
    links_texto = ""
    with tab_links_prov:
        links_texto = st.text_area("Pega las URLs o enlaces de los productos (Uno por línea, ej: Bahco):", placeholder="https://www.bahco.com/cl_es/...")
        
    input_listo = False
    if imagenes_prov or links_texto.strip():
        input_listo = True

    # PROMPT MAESTRO DE EXTRACCIÓN COMERCIAL (CAMINO B)
    prompt_proveedor = """
    Analiza detalladamente los archivos o textos provistos (pueden ser una o varias capturas de pantalla de portales de proveedores, textos técnicos copiados, o información de fichas oficiales de fabricantes). 
    Tu objetivo es actuar como un analista de compras experto y extraer los datos de cada producto de forma quirúrgica para construir una cotización limpia.

    ⚠️ REGLAS INQUEBRANTABLES DE EXTRACCIÓN:
    1. CÓDIGO / SKU: Identifica el código de fábrica, número de parte o SKU del producto. No inventes caracteres.
    2. MARCA: Identifica la marca del producto (ej. BAHCO, YATO, etc.). Si en la captura aparece el logotipo visual de la marca, interprétalo correctamente.
    3. DESCRIPCIÓN TÉCNICA: Extrae el nombre del producto and sus especificaciones clave (medidas, capacidad en toneladas, voltajes, etc.) de forma de título ejecutivo y limpio en español.
    4. DETECCIÓN DE COSTO NETO:
       - SI ES UN PANTALLAZO DE PORTAL DE PROVEEDOR: Busca el valor numérico que corresponda a tu costo de adquisición neto (puede figurar como "Precio Distribuidor", "Costo Neto", "Mi Precio", "Precio Mayorista"). Extráelo como un número entero o flotante puro sin puntos ni signos de peso.
       - SI ES UN ENLACE DE FABRICANTE O NO TIENE PRECIO VISIBLE (Caso Catálogos tipo Bahco): No intentes buscar o inventar un valor. Asigna ESTRICTAMENTE el valor 0.0 en el campo de costo.

    Devuelve ÚNICAMENTE un JSON puro con esta estructura, sin bloques markdown (```json):
    {
        "productos": [
            {"codigo": "CÓDIGO", "marca": "MARCA", "descripcion": "DESCRIPCIÓN", "costo_neto": 0.0, "cantidad": 1}
        ]
    }
    """

# PROCESADOR MAESTRO AL HACER CLIC EN GENERAR
if input_listo and api_key:
    if st.button("🔥 Generar Cotización Inteligente"):
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-2.5-flash')
            cotizacion_final = []
            
            # --- RUTA DE EJECUCIÓN A: CATÁLOGO INTERNO ---
            if modo_operacion == "Catálogo Interno (Precio Lista)":
                st.info("🔄 Buscando equivalencias y cruzando catálogo...")
                df = df_catalogo.copy()
                df.columns = [str(c).strip() for c in df.columns]
                columnas_disponibles = list(df.columns)
                n_cols = len(columnas_disponibles)
                
                idx_cod = next((i for i, c in enumerate(columnas_disponibles) if 'cod' in c.lower() or 'id' in c.lower()), 0)
                idx_desc = next((i for i, c in enumerate(columnas_disponibles) if 'desc' in c.lower() or 'nom' in c.lower() or 'art' in c.lower() or 'prod' in c.lower() or 'det' in c.lower()), 1)
                idx_precio = next((i for i, c in enumerate(columnas_disponibles) if 'prec' in c.lower() or 'val' in c.lower() or 'neto' in c.lower() or 'unit' in c.lower()), n_cols - 1)
                idx_marca = next((i for i, c in enumerate(columnas_disponibles) if 'mar' in c.lower() or 'bra' in c.lower() or 'fab' in c.lower()), -1)
                
                col_codigo, col_desc, col_precio = columnas_disponibles[idx_cod], columnas_disponibles[idx_desc], columnas_disponibles[idx_precio]
                col_marca = columnas_disponibles[idx_marca] if idx_marca != -1 else None
                
                df['__desc_clean'] = df[col_desc].apply(normalizar_texto)
                df['__cod_clean'] = df[col_codigo].apply(normalizar_texto)
                dict_equivalencias = leer_equivalencias("equivalencias.csv")
                
                response = model.generate_content(contenido_para_gemini)
                datos_pedido = json.loads(response.text.strip().replace("```json", "").replace("```", ""))
                lista_productos = datos_pedido.get("productos", [])
                cantidades_dict = {item.get("busqueda", ""): int(item.get("cantidad", 1)) for item in lista_productos}
                candidates_rag = {}
                
                for item in lista_productos:
                    termino = item.get("busqueda", "")
                    if not termino: continue
                    norm_termino = normalizar_texto(termino)
                    codigo_forzado = dict_equivalencias.get(norm_termino, None)
                    if not codigo_forzado:
                        for k_eq, v_eq in dict_equivalencias.items():
                            if k_eq in norm_termino or norm_termino in k_eq:
                                codigo_forzado = v_eq; break
                    if codigo_forzado:
                        match_rows = df[df['__cod_clean'] == normalizar_texto(codigo_forzado)]
                        if not match_rows.empty:
                            candidates_rag[termino] = [{"codigo": str(match_rows.iloc[0][col_codigo]), "descripcion": str(match_rows.iloc[0][col_desc]), "precio": float(limpiar_precio(match_rows.iloc[0][col_precio]))}]
                            continue
                    palabras_flat = set()
                    for t in [termino] + item.get("sinonimos", []):
                        for p in limpiar_plurales(normalizar_texto(t)).split():
                            if len(p) > 2 and p not in STOP_WORDS: palabras_flat.add(p)
                    if not palabras_flat: palabras_flat = {normalizar_texto(termino)}
                    df['__tmp_score'] = df.apply(lambda r: sum(5 for p in palabras_flat if p in " " + str(r['__desc_clean']) + " " + str(r['__cod_clean']) + " "), axis=1)
                    df_filtrado = df[df['__tmp_score'] > 0].sort_values(by='__tmp_score', ascending=False).head(40)
                    cand_list = []
                    for _, r in df_filtrado.iterrows():
                        cand_list.append({"codigo": str(r[col_codigo]), "descripcion": str(r[col_desc]), "precio": float(limpiar_precio(r[col_precio]))})
                    candidates_rag[termino] = cand_list
                
                prompt_resolucion = "Actúas como un experto. Mapea el catálogo:\n" + json.dumps(candidates_rag, ensure_ascii=False) + "\nDevuelve JSON con resultados:[{busqueda_original, codigo_elegido, descripcion_elegida, precio_elegido}]."
                response_res = model.generate_content(prompt_resolucion)
                datos_finales = json.loads(response_res.text.strip().replace("```json", "").replace("```", ""))
                
                for res in datos_finales.get("resultados", []):
                    origen = res.get("busqueda_original", "")
                    cant = cantidades_dict.get(origen, 1)
                    cod = str(res.get("codigo_elegido", "MANUAL")).strip()
                    desc = str(res.get("descripcion_elegida", "❌ NO ENCONTRADO"))
                    px_lista = float(res.get("precio_elegido", 0.0))
                    marca = "YATO"
                    if cod != "MANUAL":
                        match_rows = df[df['__cod_clean'] == normalizar_texto(cod)]
                        if not match_rows.empty:
                            desc = str(match_rows.iloc[0][col_desc])
                            px_lista = float(limpiar_precio(match_rows.iloc[0][col_precio]))
                            if col_marca: marca = str(match_rows.iloc[0][col_marca]).upper()
                    
                    px_final_neto = px_lista - (px_lista * (descuento_aplicar / 100)) if descuento_aplicar > 0 else px_lista
                    pm_val = limpiar_precio(precio_manual_input)
                    if pm_val > 0: px_final_neto = pm_val
                    
                    cotizacion_final.append({
                        "Código": cod, "Marca": marca, "Descripción Catálogo": desc, "Cantidad": cant,
                        "Precio Lista (Neto)": px_lista, "Descuento Aplicado": f"{descuento_aplicar}%" if descuento_aplicar > 0 else "0%",
                        "Precio Final (Neto)": px_final_neto, "Total Neto": px_final_neto * cant
                    })

            # --- RUTA DE EJECUCIÓN B: COSTO PROVEEDOR + MARGEN ---
            else:
                st.info("🔄 Analizando capturas y extrayendo fichas técnicas web...")
                items_extraidos = []
                
                if imagenes_prov:
                    for img_file in imagenes_prov:
                        response_img = model.generate_content([prompt_proveedor, Image.open(img_file)])
                        try:
                            json_data = json.loads(response_img.text.strip().replace("```json", "").replace("```", ""))
                            items_extraidos.extend(json_data.get("productos", []))
                        except: pass
                        
                if links_texto.strip():
                    urls = [u.strip() for u in links_texto.split("\n") if u.strip()]
                    texto_consolidado_urls = ""
                    for u in urls:
                        texto_consolidado_urls += f"\n--- CONTENIDO TÉCNICO DE LA URL: {u} ---\n" + extraer_texto_de_url(u)
                    if texto_consolidado_urls.strip():
                        response_link = model.generate_content(prompt_proveedor + "\n\nTEXTO CAPTURADO:\n" + texto_consolidado_urls)
                        try:
                            json_data = json.loads(response_link.text.strip().replace("```json", "").replace("```", ""))
                            items_extraidos.extend(json_data.get("productos", []))
                        except: pass
                
                for item in items_extraidos:
                    cod = str(item.get("codigo", "MANUAL")).strip()
                    marca = str(item.get("marca", "PROVEEDOR")).upper()
                    desc = str(item.get("descripcion", "Producto Extraído"))
                    costo_base = float(item.get("costo_neto", 0.0))
                    cant = int(item.get("cantidad", 1))
                    
                    px_final_neto = costo_base * (1 + (margen_objetivo / 100))
                    pm_val = limpiar_precio(precio_manual_input)
                    if pm_val > 0: px_final_neto = pm_val
                    
                    cotizacion_final.append({
                        "Código": cod, "Marca": marca, "Descripción Catálogo": desc, "Cantidad": cant,
                        "Precio Lista (Neto)": costo_base,
                        "Descuento Aplicado": f"+{margen_objetivo}% Margen",
                        "Precio Final (Neto)": px_final_neto, "Total Neto": px_final_neto * cant
                    })

            if cotizacion_final:
                st.session_state['df_resultado'] = pd.DataFrame(cotizacion_final)
                st.success("¡Datos procesados! Revisa el cuadro interactivo abajo.")
        except Exception as e:
            st.error(f"Error en procesamiento comercial: {e}")

# RENDERIZADO INTERACTIVO MAESTRO DESDE MEMORIA (Saneado definitivo)
if st.session_state['df_resultado'] is not None:
    st.markdown("### 📱 Cuadro Comercial Express (Editable en Pantalla)")
    st.caption("💡 Truco Comercial: Si algún producto de un link viene con costo $0, puedes hacer doble clic en la celda 'Precio Lista (Neto)', digitar el valor real, presionar Enter y los cálculos se actualizarán al instante.")
    
    # SOLUCIÓN DE SANEAMIENTO: Las columnas con diseño visual (Precio Lista, Precio Final, Total Neto)
    # se omiten POR COMPLETO de la lista global de 'disabled'. Así no existe ningún cruce de variables.
    df_editable = st.data_editor(
        st.session_state['df_resultado'],
        column_config={
            "Precio Lista (Neto)": st.column_config.NumberColumn("Costo Base / Lista ($)", format="$%.0f"),
            "Precio Final (Neto)": st.column_config.NumberColumn("P. Venta Neto ($)", format="$%.0f"),
            "Total Neto": st.column_config.NumberColumn("Total Neto ($)", format="$%.0f"),
            "Cantidad": st.column_config.NumberColumn("Cant", min_value=1),
        },
        disabled=["Código", "Marca", "Descripción Catálogo", "Descuento Aplicado"],
        use_container_width=True
    ).copy()
    
    # Recalculador dinámico en tiempo real basado en las modificaciones en pantalla
    pm_global = limpiar_precio(precio_manual_input)
    if modo_operacion == "Catálogo Interno (Precio Lista)":
        if pm_global > 0: df_editable["Precio Final (Neto)"] = pm_global
        else: df_editable["Precio Final (Neto)"] = df_editable["Precio Lista (Neto)"] * (1 - (descuento_aplicar / 100))
    else:
        if pm_global > 0: df_editable["Precio Final (Neto)"] = pm_global
        else: df_editable["Precio Final (Neto)"] = df_editable["Precio Lista (Neto)"] * (1 + (margen_objetivo / 100))
        
    df_editable["Total Neto"] = df_editable["Precio Final (Neto)"] * df_editable["Cantidad"]
    
    # Actualización de Métricas Globales
    subtotal_lista_v = sum(df_editable["Precio Lista (Neto)"] * df_editable["Cantidad"])
    total_neto_final_v = sum(df_editable["Total Neto"])
    descuento_total_pesos_v = max(subtotal_lista_v - total_neto_final_v, 0.0) if modo_operacion == "Catálogo Interno (Precio Lista)" else 0.0
    iva_calculado_v = total_neto_final_v * 0.19
    total_bruto_v = total_neto_final_v + iva_calculado_v
    
    st.markdown("### 📊 Desglose Económico Actualizado")
    c_neto, c_desc, c_neto_f, c_iva, c_bruto = st.columns(5)
    c_neto.metric(label="Total Base Neto", value=f"${subtotal_lista_v:,.0f}")
    c_desc.metric(label="Descuento Otorgado" if modo_operacion == "Catálogo Interno (Precio Lista)" else "Margen Aplicado", value=f"-${descuento_total_pesos_v:,.0f}" if modo_operacion == "Catálogo Interno (Precio Lista)" else f"+{margen_objetivo}%")
    c_neto_f.metric(label="Neto Final Cliente", value=f"${total_neto_final_v:,.0f}")
    c_iva.metric(label="IVA (19%)", value=f"${iva_calculado_v:,.0f}")
    c_bruto.metric(label="Total Bruto Final", value=f"${total_bruto_v:,.0f}")
    
    st.markdown("---")
    
    # Generación de archivos binarios finales
    dict_img = {}
    excel_bin = generar_excel_comercial(
        df_editable, nombre_cliente, empresa_cliente, numero_folio,
        total_neto_final_v, iva_calculado_v, total_bruto_v,
        logo_bytes, dict_img, condicion_pago_input, vendedor_input
    )
    pdf_raw = generar_pdf_comercial(
        df_editable, nombre_cliente, empresa_cliente, numero_folio,
        total_neto_final_v, iva_calculado_v, total_bruto_v,
        logo_bytes, condicion_pago_input, vendedor_input
    )
    pdf_bin = pdf_raw.encode('latin-1') if isinstance(pdf_raw, str) else bytes(pdf_raw)
    
    c_down1, c_down2 = st.columns(2)
    with c_down1:
        st.download_button(
            label="🟢 Descargar Documento Excel Premium (.xlsx)", data=excel_bin,
            file_name=f"Cotizacion_{numero_folio}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True
        )
    with c_down2:
        st.download_button(
            label="🔴 Descargar Documento PDF Oficial (.pdf)", data=pdf_bin,
            file_name=f"Cotizacion_{numero_folio}.pdf", mime="application/pdf", use_container_width=True
        )
else:
    st.info("Introduce una solicitud o pantallazo arriba y presiona Generar para activar los paneles comerciales.")
